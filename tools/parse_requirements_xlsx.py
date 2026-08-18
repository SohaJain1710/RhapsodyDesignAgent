"""
REQUIREMENTS XLSX PARSER
========================
Parses a Software Requirements Excel (.xlsx) file and extracts a structured
JSON suitable for AI-driven design generation.

Required columns (case-insensitive, partial-match tolerant):
  - ID
  - Type
  - Requirement
  - State  (also matches "State*", "Status")
  - Parametrizable
  - RealizedByModule  (also matches "Realized By Module", "Realized_By_Module")

Output JSON schema:
{
  "source"     : "filename.xlsx",
  "sheet"      : "SheetName",
  "parsed_at"  : "ISO-8601 timestamp",
  "total"      : <int>,
  "filtered"   : <int>,
  "requirements": [
    {
      "id"               : "REQ-001",
      "type"             : "Functional",
      "requirement"      : "The system shall ...",
      "state"            : "Approved",
      "parametrizable"   : "Yes",
      "realized_by_module": "rb_wimi_CustAirbagWarningLamp"
    },
    ...
  ]
}

Usage:
  python tools/parse_requirements_xlsx.py --input requirements.xlsx
  python tools/parse_requirements_xlsx.py --input requirements.xlsx --output reqs.json
  python tools/parse_requirements_xlsx.py --input requirements.xlsx --sheet "SW Reqs" --state Approved
  python tools/parse_requirements_xlsx.py --input requirements.xlsx --state-column "State_SSG"
  python tools/parse_requirements_xlsx.py --input requirements.xlsx --state-column "State_WL" --state approved
  python tools/parse_requirements_xlsx.py --input requirements.xlsx --list-sheets
"""

import os
import sys
import json
import argparse
import re
from datetime import datetime, timezone

# ── Dependency bootstrap ───────────────────────────────────────────────────────

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("⚠  openpyxl not found — attempting install...")
        import subprocess
        result = subprocess.run(
            [sys.executable, "-m", "pip", "install", "openpyxl"],
            capture_output=True, text=True
        )
        if result.returncode != 0:
            print("❌ Could not install openpyxl automatically.")
            print("   Run:  pip install openpyxl")
            sys.exit(1)
        print("✅ openpyxl installed.")
        import importlib
        importlib.invalidate_caches()

_ensure_openpyxl()
import openpyxl  # noqa: E402  (after bootstrap)


# ── Column matching ────────────────────────────────────────────────────────────

# Canonical key  →  list of patterns that should match (regex, case-insensitive)
# Default patterns — can be overridden at runtime
DEFAULT_COLUMN_PATTERNS: dict[str, list[str]] = {
    "id":                [r"^id$", r"^req.*id$", r"^identifier$"],
    "type":              [r"^type$", r"^req.*type$", r"^requirement.*type$"],
    "requirement":       [r"^requirement[s]?$", r"^req(uirement)?[ _-]?text$",
                          r"^description$", r"^req$"],
    "state":             [r"^state[*]?$", r"^status$", r"^req.*state$"],
    "state_ssg":         [r"^state.ssg$", r"^statessg$", r"^state_ssg$"],
    "parametrizable":    [r"^parametriz[ae]ble$", r"^param[_\s]?izable$",
                          r"^is.?param"],
    "realized_by_module":[r"^realizedbymodule$", r"^realized.by.module$",
                          r"^realized_by_module$", r"^module$",
                          r"^implementing.module$"],
}

# Runtime patterns (can be modified via CLI or programmatically)
COLUMN_PATTERNS = DEFAULT_COLUMN_PATTERNS.copy()


def _normalize(text: str) -> str:
    """Lowercase, collapse whitespace/underscores/hyphens for matching."""
    return re.sub(r"[\s_\-]+", "", str(text).lower().strip())


def _match_column(header: str) -> str | None:
    """Return canonical key if header matches any pattern, else None."""
    norm = _normalize(header)
    for key, patterns in COLUMN_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, norm):
                return key
    return None


def _map_headers(row) -> dict[str, int]:
    """
    Given the header row cells, return {canonical_key: col_index (0-based)}.
    Reports unmapped canonical keys as warnings.
    """
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell.value is None:
            continue
        key = _match_column(str(cell.value))
        if key and key not in mapping:
            mapping[key] = idx

    missing = [k for k in COLUMN_PATTERNS if k not in mapping]
    if missing:
        print(f"⚠  Could not map columns: {missing}")
        print("   Available headers:", [c.value for c in row if c.value])
    return mapping


# ── Sheet selection ────────────────────────────────────────────────────────────

def _pick_sheet(wb: "openpyxl.Workbook", name: str | None):
    """Return the worksheet to process, or raise ValueError."""
    if name:
        if name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{name}' not found. Available: {wb.sheetnames}"
            )
        return wb[name]
    # Auto-pick: prefer a sheet whose name contains 'req' (case-insensitive)
    for sname in wb.sheetnames:
        if "req" in sname.lower():
            print(f"ℹ  Auto-selected sheet: '{sname}'")
            return wb[sname]
    # Fall back to first sheet
    print(f"ℹ  Using first sheet: '{wb.sheetnames[0]}'")
    return wb.active


# ── Header row detection ───────────────────────────────────────────────────────

def _find_header_row(ws, max_scan: int = 20) -> int:
    """
    Scan up to max_scan rows for the row that best matches required column names.
    Returns 1-based row index.
    """
    best_row  = 1
    best_score = 0
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row    = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        mapped = _map_headers(row)
        score  = len(mapped)
        if score > best_score:
            best_score = score
            best_row   = row_idx
    return best_row


# ── Cell value normalisation ───────────────────────────────────────────────────

def _cell_str(cell) -> str:
    """Return clean string representation of a cell value."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


# ── Core parser ───────────────────────────────────────────────────────────────

def parse_requirements(
    xlsx_path: str,
    sheet_name: str | None = None,
    state_filter: list[str] | None = None,
    skip_empty_id: bool = True,
    state_patterns: list[str] | None = None,
    state_column: str | None = None,
) -> dict:
    """
    Parse an xlsx requirements file and return a structured dict.

    Parameters
    ----------
    xlsx_path    : path to the .xlsx file
    sheet_name   : explicit sheet name (None = auto-detect)
    state_filter : if set, keep only rows whose state matches (case-insensitive)
    skip_empty_id: skip rows where ID cell is empty (default True)
    state_patterns: custom regex patterns to match state column headers
                    (None = use defaults, e.g. ["^state_.*$", "^custom_status$"])
    state_column : explicit column name to use as state (overrides pattern matching)
                   (e.g. "State_SSG", "State_WL")
    """
    # Apply custom state patterns if provided
    global COLUMN_PATTERNS
    if state_patterns is not None:
        COLUMN_PATTERNS = DEFAULT_COLUMN_PATTERNS.copy()
        COLUMN_PATTERNS["state"] = state_patterns
    else:
        COLUMN_PATTERNS = DEFAULT_COLUMN_PATTERNS.copy()
    
    # If explicit state_column is provided, we'll use it directly later
    explicit_state_col = state_column

    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = _pick_sheet(wb, sheet_name)

    # ── Suppress header-scanning warnings; redo cleanly after detecting row ──
    import io, contextlib
    header_row_idx = _find_header_row(ws, max_scan=20)

    rows      = list(ws.iter_rows(min_row=header_row_idx))
    header    = rows[0]
    
    # If explicit state_column provided, find its index and add to col_map
    if explicit_state_col:
        explicit_state_idx = None
        for idx, cell in enumerate(header):
            if cell.value and str(cell.value).strip() == explicit_state_col:
                explicit_state_idx = idx
                break
        if explicit_state_idx is None:
            raise ValueError(
                f"State column '{explicit_state_col}' not found in sheet. "
                f"Available columns: {[c.value for c in header if c.value]}"
            )
    
    col_map   = _map_headers(header)
    
    # Override state mapping if explicit column provided
    if explicit_state_col is not None:
        col_map["state"] = explicit_state_idx

    if not col_map:
        raise ValueError("Could not identify any required columns in this sheet.")

    # normalise state_filter
    state_set = {s.lower().strip() for s in state_filter} if state_filter else None

    requirements = []
    skipped_rows = 0

    for data_row in rows[1:]:
        def get(key: str) -> str:
            idx = col_map.get(key)
            return _cell_str(data_row[idx]) if idx is not None else ""

        req_id = get("id")

        # Skip entirely empty rows
        if all(_cell_str(c) == "" for c in data_row):
            continue

        if skip_empty_id and not req_id:
            skipped_rows += 1
            continue

        # State filtering only applies if 'state' column exists
        if "state" in col_map:
            state = get("state")
            if state_set and state.lower().strip() not in state_set:
                skipped_rows += 1
                continue

        entry = {
            "id"               : req_id,
            "type"             : get("type"),
            "requirement"      : get("requirement"),
            "parametrizable"   : get("parametrizable"),
            "realized_by_module": get("realized_by_module"),
        }
        
        # Only include state if the column exists
        if "state" in col_map:
            entry["state"] = get("state")
        
        # Include state_ssg if the column exists
        if "state_ssg" in col_map:
            entry["state_ssg"] = get("state_ssg")
        
        requirements.append(entry)

    wb.close()

    result = {
        "source"      : os.path.basename(xlsx_path),
        "sheet"       : ws.title,
        "parsed_at"   : datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "total_rows"  : len(rows) - 1,
        "skipped_rows": skipped_rows,
        "total"       : len(requirements),
        "state_filter": state_filter or [],
        "column_map"  : {k: header[v].value for k, v in col_map.items()},
        "requirements": requirements,
    }
    return result


# ── Summary helpers ────────────────────────────────────────────────────────────

def _print_summary(data: dict) -> None:
    print("\n─── Parse Summary ───────────────────────────────")
    print(f"  Source      : {data['source']}")
    print(f"  Sheet       : {data['sheet']}")
    print(f"  Total rows  : {data['total_rows']}")
    print(f"  Skipped     : {data['skipped_rows']}")
    print(f"  Extracted   : {data['total']}")

    if data["column_map"]:
        print(f"\n  Column mapping:")
        for key, header in data["column_map"].items():
            print(f"    {key:<22} ← \"{header}\"")

    # Group by type
    from collections import Counter
    types     = Counter(r["type"]  for r in data["requirements"] if r["type"])
    states    = Counter(r.get("state")  for r in data["requirements"] if r.get("state", ""))
    states_sg = Counter(r.get("state_ssg") for r in data["requirements"] if r.get("state_ssg", ""))
    if types:
        print(f"\n  Requirement types : {dict(types)}")
    if states:
        print(f"  States            : {dict(states)}")
    if states_sg:
        print(f"  State_SSG         : {dict(states_sg)}")

    # Unique modules
    modules = sorted({r["realized_by_module"]
                      for r in data["requirements"] if r["realized_by_module"]})
    if modules:
        print(f"\n  Modules ({len(modules)}):")
        for m in modules:
            print(f"    • {m}")
    
    # ── Summary Table ──
    print("\n  ╔════════════════════════════════════════════════════════════╗")
    print("  ║            REQUIREMENT SUMMARY TABLE                      ║")
    print("  ╠════════════════════════════════════════════════════════════╣")
    
    total_count = data["total"]
    rqmt_approved = sum(1 for r in data["requirements"] 
                       if r.get("type", "").strip() == "Rqmt" 
                       and r.get("state", "").lower().strip() == "approved")
    rqmt_implemented = sum(1 for r in data["requirements"] 
                          if r.get("type", "").strip() == "Rqmt" 
                          and r.get("state", "").lower().strip() == "implemented")
    
    print(f"  │ Total Requirements                            : {total_count:>8} │")
    print("  ├────────────────────────────────────────────────────────────┤")
    print(f"  │ Type=Rqmt, State=Approved                    : {rqmt_approved:>8} │")
    print(f"  │ Type=Rqmt, State=Implemented                 : {rqmt_implemented:>8} │")
    print("  ├────────────────────────────────────────────────────────────┤")
    print("  │ REQUIREMENT TYPE DISTRIBUTION:                             │")
    
    # Add type distribution
    for req_type, count in sorted(types.items(), key=lambda x: x[1], reverse=True):
        print(f"  │   {req_type:<40} : {count:>8} │")
    
    print("  ╚════════════════════════════════════════════════════════════╝")
    
    print("─────────────────────────────────────────────────\n")


# ── CLI ────────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(
        description="Parse a Software Requirements .xlsx into structured JSON.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--input",  "-i", required=True,
                    help="Path to the .xlsx requirements file")
    ap.add_argument("--output", "-o", default=None,
                    help="Path for the output .json file "
                         "(default: <input_stem>_requirements.json)")
    ap.add_argument("--sheet",  "-s", default=None,
                    help="Sheet name to parse (default: auto-detect)")
    ap.add_argument("--state",  "-f", nargs="*", default=None,
                    metavar="STATE",
                    help="Filter by state value(s), e.g. --state Approved Valid "
                         "(default: include all states)")
    ap.add_argument("--state-column", default=None,
                    metavar="COLUMN_NAME",
                    help="Explicit column name to use as state attribute "
                         "(e.g. --state-column 'State_SSG'). "
                         "Overrides pattern matching.")
    ap.add_argument("--state-patterns", nargs="+", default=None,
                    metavar="PATTERN",
                    help="Custom regex patterns to match state column headers "
                         "(e.g. --state-patterns '^state_.*$' '^custom_status$'). "
                         "Default: ^state[*]?$, ^status$, ^req.*state$")
    ap.add_argument("--list-sheets", action="store_true",
                    help="List all sheet names and exit")
    ap.add_argument("--no-skip-empty-id", action="store_true",
                    help="Include rows where ID cell is empty")
    return ap


def main(argv: list[str] | None = None) -> None:
    ap  = build_parser()
    args = ap.parse_args(argv)

    # ── List sheets mode ──
    if args.list_sheets:
        wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
        print(f"\nSheets in '{os.path.basename(args.input)}':")
        for i, name in enumerate(wb.sheetnames, 1):
            ws = wb[name]
            print(f"  {i:>2}. {name}  ({ws.max_row} rows × {ws.max_column} cols)")
        wb.close()
        return

    # ── Parse ──
    try:
        data = parse_requirements(
            xlsx_path    = args.input,
            sheet_name   = args.sheet,
            state_filter = args.state,
            skip_empty_id= not args.no_skip_empty_id,
            state_patterns = args.state_patterns,
            state_column = args.state_column,
        )
    except (FileNotFoundError, ValueError) as exc:
        print(f"Error: {exc}")
        sys.exit(1)

    _print_summary(data)

    # ── Write JSON ──
    if args.output:
        out_path = args.output
    else:
        # Extract module name from realized_by_module (use first non-empty one)
        module_name = None
        for req in data["requirements"]:
            if req.get("realized_by_module", "").strip():
                module_name = req["realized_by_module"].strip()
                break
        
        # If no module found, use input stem
        if not module_name:
            module_name = os.path.splitext(os.path.basename(args.input))[0]
        
        out_dir  = os.path.dirname(os.path.abspath(args.input))
        out_path = os.path.join(out_dir, f"{module_name}ParsedReq.json")

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"✅ JSON written → {out_path}")
    print(f"   {data['total']} requirements ready for AI design generation.\n")


# ── Library entry-point ────────────────────────────────────────────────────────

def load_requirements_json(xlsx_path: str, **kwargs) -> dict:
    """
    Convenience function for importing in other scripts:

        from tools.parse_requirements_xlsx import load_requirements_json
        data = load_requirements_json("requirements.xlsx", state_filter=["Approved"])
        data = load_requirements_json("requirements.xlsx", 
                                      state_patterns=["^state_.*$", "^workflow_status$"])
        data = load_requirements_json("requirements.xlsx", state_column="State_SSG")
    """
    return parse_requirements(xlsx_path, **kwargs)


if __name__ == "__main__":
    main()
